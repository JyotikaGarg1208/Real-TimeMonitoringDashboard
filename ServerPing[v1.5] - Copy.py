import datetime
import os
import openpyxl
from openpyxl.styles import PatternFill
from datetime import date
import calendar
import time
import shutil
import win32com.client
import tkinter as tk
from PIL import Image, ImageTk, ImageEnhance, ImageDraw, ImageFont
import aioping
import asyncio
import warnings
import socket
from tkinter import simpledialog, messagebox, ttk
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import matplotlib.pyplot as plt
from tkinter import Toplevel
from winotify import Notification, audio
import random
import math
import re
import textwrap

# Suppress OpenPyXL warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

def extract_text_with_brackets(text):
    """Extract all text inside parentheses and return a list of parts."""
    parts = re.split(r"(\(.*?\))", text)  # Split by parentheses
    return parts 

class PingingDashboard:
    def __init__(self, root):
        self.root = root
        self.device_type = None
        self.ask_device_type()
        if self.device_type:
            print(f"Device Type Selected: {self.device_type}")

        self.server_type = None
        self.function_type = None
        self.os_type = None
        self.current_page_index = 0
        self.cycle_count = 0  # Counter for completed rounds
        self.current_ping_cycle = 0
        self.categories = ["Active", "Delayed", "Unresponsive"]
        self.auto_refresh = True
        self.max_cards_per_page = 25
        self.card_height = 180
        self.card_width = 350
        self.ping_interval_manually_changed = False
        self.pages = {category: [] for category in self.categories}
        self.cycle_interval = 5000  # Auto-refresh every 5 seconds for page cycling
        self.servers = []  # Stores all servers for searching and sorting
        self.filtered_servers = []  # Stores filtered servers for searching
        self.sort_by = "Server Name"  # Default sort option
        self.search_active = False  # To track if search is active
        self.server_data = {}  # Initialize server data for graph
        self.response_times = {}
        self.notified_servers = {"unresponsive": set(), "delayed": set(), "active": set()}  # Track notified servers
        self.initial_notification_done = False  # Flag to track the initial run
        self.ping_interval = None
        self.ping_in_progress = False  # Add this flag to track ongoing ping cycles
        self.auto_refresh_id = None  # Auto-refresh loop identifier
        self.auto_refresh_var = tk.BooleanVar(value=True)
        self.search_var = tk.StringVar()  # Initialize search_var for the search field
        self.sort_requested = False
        self.card_clicked = False
        self.flip_animation_running = False
        self.auto_flipped = False
        self.ping_interval_var = tk.IntVar(value=10)  # Default ping interval to 10 seconds
        self.header_row = None  # If not set by the user, we fallback to row=8

        self.setup_ui()

    def ask_device_type(self):
        """Dialog to ask for device type input."""
        dialog = tk.Toplevel(self.root)
        dialog.geometry("300x170")
        dialog.title("Device Type")
        dialog.configure(bg='#2b2b2b')
        dialog.resizable(False, False)

        def set_device_type():
            input_value = device_entry.get().strip()
            if input_value:
                self.device_type = input_value.capitalize()
            else:
                self.device_type = "Server"
            dialog.destroy()

        tk.Label(dialog, text="Enter the Monitoring Device Type", font=("Segoe UI", 12), fg="white", bg='#2b2b2b').pack(pady=20)
        device_entry = ttk.Entry(dialog, width=20, font=("Segoe UI", 12))
        device_entry.pack(pady=7, padx=20)
        ok_button = tk.Button(dialog, text="OK", command=set_device_type, bg='#74a8e7', fg="white", height=3, width=22, font=("Segoe UI", 11), relief="flat")
        ok_button.pack(pady=10)
        dialog.grab_set()
        self.root.wait_window(dialog)

    def start_scan(self):
        """Enhanced dialog boxes for Scan IPs input."""
        self.server_data = {}

        def ask_filename():
            dialog = tk.Toplevel(self.root)
            dialog.geometry("300x170")
            dialog.title("File Name")
            dialog.configure(bg='#2b2b2b')
            dialog.resizable(False, False)

            def set_filename():
                file_name = file_entry.get()
                if file_name and os.path.isfile(file_name):
                    dialog.destroy()
                    # Ask the user for which row contains headers
                    self.ask_header_row(file_name)
                else:
                    messagebox.showerror("Error", "File not found. Please enter a valid file name.")

            tk.Label(dialog, text="Enter the Excel file name", font=("Segoe UI", 12), fg="white", bg='#2b2b2b').pack(pady=20)
            file_entry = ttk.Entry(dialog, font=("Segoe UI", 12))
            file_entry.pack(pady=7, padx=20)
            ok_button = tk.Button(dialog, text="OK", command=set_filename, bg='#74a8e7', height=3, width=22, fg="white", font=("Segoe UI", 11), relief="flat")
            ok_button.pack(pady=10)
            dialog.grab_set()
            self.root.wait_window(dialog)

        def ask_header_row(file_name):
            """Ask which row has the headers, store in self.header_row, then proceed."""
            dialog = tk.Toplevel(self.root)
            dialog.geometry("300x170")
            dialog.title("Header Row")
            dialog.configure(bg='#2b2b2b')
            dialog.resizable(False, False)

            tk.Label(dialog, text="Enter the row number containing headers", font=("Segoe UI", 12), fg="white", bg='#2b2b2b').pack(pady=20)
            header_row_var = tk.IntVar(value=8)  # default to row=8 if you like
            header_entry = ttk.Entry(dialog, textvariable=header_row_var, font=("Segoe UI", 12))
            header_entry.pack(pady=7, padx=20)

            def confirm_header():
                value = header_row_var.get()
                value = value + 1
                if value < 1:
                    messagebox.showerror("Error", "Header row must be >= 1.")
                else:
                    self.header_row = value  # store user selection
                    dialog.destroy()
                    # Next: select columns
                    self.select_columns(file_name)

            ok_btn = tk.Button(dialog, text="OK", command=confirm_header, bg='#74a8e7', fg="white", height=2, width=22, font=("Segoe UI", 11), relief="flat")
            ok_btn.pack(pady=10)
            dialog.grab_set()
            self.root.wait_window(dialog)

        # Expose ask_header_row for usage
        self.ask_header_row = ask_header_row
        ### ADDITIONS END

        ask_filename()

    def select_columns(self, file_name):
        """Allow the user to select columns from the Excel headers."""
        dialog = tk.Toplevel(self.root)
        dialog.geometry("300x320")
        dialog.title("Select Columns")
        dialog.configure(bg='#2b2b2b')
        dialog.resizable(False, False)

        wb = openpyxl.load_workbook(filename=file_name, read_only=True)
        ws = wb.active

        ### ADDITIONS START
        # If user didn't set self.header_row, fallback to 8
        row_for_headers = (self.header_row - 1) if self.header_row else 8
        ### ADDITIONS END

        # Read the selected row for headers
        headers = [cell.value for cell in ws[row_for_headers]]
        wb.close()

        # Now create variables to hold the selections
        server_column_var = tk.StringVar()
        ip_column_var = tk.StringVar()
        dashboard_column_var = tk.StringVar()

        tk.Label(dialog, text=f"Select {self.device_type} Name Column", font=("Segoe UI", 12), fg="white", bg='#2b2b2b').pack(pady=10)
        server_column_menu = ttk.Combobox(dialog, textvariable=server_column_var, values=headers, font=("Segoe UI", 10))
        server_column_menu.pack(pady=1, padx=20)

        tk.Label(dialog, text="Select IP Address Column", font=("Segoe UI", 12), fg="white", bg='#2b2b2b').pack(pady=10)
        ip_column_menu = ttk.Combobox(dialog, textvariable=ip_column_var, values=headers, font=("Segoe UI", 10))
        ip_column_menu.pack(pady=1, padx=20)

        tk.Label(dialog, text="Select Dashboard Column (optional)", font=("Segoe UI", 12), fg="white", bg='#2b2b2b').pack(pady=10)
        dashboard_column_menu = ttk.Combobox(dialog, textvariable=dashboard_column_var, values=headers, font=("Segoe UI", 10))
        dashboard_column_menu.pack(pady=1, padx=20)

        def select_additional_columns():
            server_column = server_column_var.get()
            ip_column = ip_column_var.get()
            dashboard_column = dashboard_column_var.get() or None  # Optional

            if server_column and ip_column:
                dialog.destroy()
                self.select_additional_columns(file_name, headers, [server_column, ip_column, dashboard_column])
            else:
                messagebox.showerror("Error", "Please select the required columns.")

        ok_button = tk.Button(dialog, text="Next", command=select_additional_columns, bg='#74a8e7', height=1, width=22, fg="white", font=("Segoe UI", 10), relief="flat")
        ok_button.pack(pady=20)

        dialog.grab_set()
        self.root.wait_window(dialog)

    def select_additional_columns(self, file_name, headers, selected_columns):
        """Allow the user to select any three columns to display on the back of the card."""
        dialog = tk.Toplevel(self.root)
        dialog.geometry("300x350")
        dialog.title("Select Columns to Display")
        dialog.configure(bg='#2b2b2b')
        dialog.resizable(False, False)

        # Exclude already selected columns
        available_headers = [h for h in headers if h not in selected_columns and h is not None]

        # Variables to hold the selected columns
        additional_columns_vars = [tk.StringVar(), tk.StringVar(), tk.StringVar()]

        tk.Label(dialog, text="Select Columns to Display", font=("Segoe UI", 12), fg="white", bg='#2b2b2b', wraplength=380).pack(pady=10)

        for i in range(3):
            tk.Label(dialog, text=f"Select Column {i+1}", font=("Segoe UI", 12), fg="white", bg='#2b2b2b').pack(pady=10)
            column_menu = ttk.Combobox(dialog, textvariable=additional_columns_vars[i], values=available_headers, font=("Segoe UI", 10))
            column_menu.pack(pady=1, padx=20)

        def set_columns():
            selected_additional_columns = [var.get() for var in additional_columns_vars]
            if all(selected_additional_columns):
                dialog.destroy()
                data = self.read_excel_with_headers(
                    file_name,
                    selected_columns[0],  # Server Name Column
                    selected_columns[1],  # IP Address Column
                    selected_columns[2],  # Dashboard? Column
                    selected_additional_columns
                )
                self.start_ping(data)  # Initial ping
                self.categorize_servers(data)
                self.cycle_pages()
                self.rerun_aioping(reset_cycle=True)
            else:
                messagebox.showerror("Error", "Please select all three columns.")

        ok_button = tk.Button(dialog, text="OK", command=set_columns, bg='#74a8e7', height=1, width=22, fg="white", font=("Segoe UI", 10), relief="flat")
        ok_button.pack(pady=20)

        dialog.grab_set()
        self.root.wait_window(dialog)
        
    def setup_ui(self):
        """Setup the UI, including ping slider and page refresh logic."""
        self.root.title(f"{self.device_type} Status Dashboard")
        self.root.configure(bg='#1e1e1e')
        self.root.geometry("1024x768")
        self.root.minsize(800, 600)  # Minimum size to prevent shrinking too small
        self.root.attributes("-fullscreen", True)

        # Title Frame
        title_frame = tk.Frame(self.root, bg='#1e1e1e')
        title_frame.pack(fill='x', padx=20, pady=10)

        # Title Label
        self.title_label = tk.Label(title_frame, text=f'7740 {self.device_type} Status Dashboard', fg="white",
                                    font=("Segoe UI", 26, "bold"), bg='#1e1e1e')
        self.title_label.pack(side=tk.LEFT)

        # Logo
        try:
            img = Image.open('dashboard.png')
            img = img.resize((200, 60))
            tkimage = ImageTk.PhotoImage(img)
            self.imagelbl = tk.Label(title_frame, image=tkimage, bg='#1e1e1e')
            self.imagelbl.image = tkimage
            self.imagelbl.pack(side=tk.RIGHT)
        except Exception as e:
            print(f"Error loading image: {e}")
            self.imagelbl = tk.Label(title_frame, text="Company ABC", fg="white",
                                     font=("Segoe UI", 18, "bold"), bg='#1e1e1e')
            self.imagelbl.pack(side=tk.RIGHT)

        # Control Frame
        control_frame = tk.Frame(self.root, bg='#e0e0e0')
        control_frame.pack(fill='x', padx=20, pady=10)

        # Buttons and controls within control frame
        self.scan_button = tk.Button(control_frame, text="SCAN IPs", fg="white", font=("Segoe UI", 12, "bold"), command=self.start_scan,
                                     bg='#74a8e7', relief='flat')
        self.scan_button.pack(side=tk.LEFT, padx=0)

        self.auto_refresh_var = tk.BooleanVar(value=True)
        self.auto_refresh_checkbox = tk.Checkbutton(control_frame, text="Auto Refresh", variable=self.auto_refresh_var, fg="black",
                                                    font=("Segoe UI", 11), bg='#e0e0e0', relief='flat', command=self.toggle_auto_refresh)
        self.auto_refresh_checkbox.pack(side=tk.LEFT, padx=5)

        # Ping Interval slider
        slider_frame = tk.Frame(control_frame, bg='#e0e0e0')
        slider_frame.pack(side=tk.LEFT, padx=10)
        self.ping_interval_label = tk.Label(slider_frame, text="Ping Interval (s):", fg="black", bg='#e0e0e0', font=("Segoe UI", 11))
        self.ping_interval_label.pack(side=tk.LEFT, padx=5)
        self.ping_interval_var = tk.IntVar(value=5)  # Default value is 5 seconds
        self.ping_interval_slider = tk.Scale(slider_frame, from_=5, to=60, orient=tk.HORIZONTAL, variable=self.ping_interval_var,
                                             bg='#e0e0e0', fg='black', font=("Segoe UI", 6), relief='flat', sliderlength=20, width=6)
        self.ping_interval_slider.pack(side=tk.LEFT)

        # Bind the slider to immediately update the ping interval
        self.ping_interval_slider.bind("<ButtonRelease-1>", self.on_slider_change)

        # Quit button
        self.quit_button = tk.Button(control_frame, text="QUIT", fg="white", font=("Segoe UI", 12, "bold"), command=self.exit_app,
                                     bg='#d32f2f', relief='flat')
        self.quit_button.pack(side=tk.RIGHT, padx=0)

        # Search and Sort
        search_frame = tk.Frame(control_frame, bg='#e0e0e0')
        search_frame.pack(side=tk.RIGHT, padx=15)
        self.search_label = tk.Label(search_frame, text="Search:", fg="black", bg='#e0e0e0', font=("Segoe UI", 11))
        self.search_label.grid(row=0, column=0, padx=5)
        self.search_var = tk.StringVar()
        self.search_entry = tk.Entry(search_frame, textvariable=self.search_var, font=("Segoe UI", 11), relief='flat', width=25)
        self.search_entry.grid(row=0, column=1, padx=0, sticky="ew")
        self.search_button = tk.Button(search_frame, command=self.search_server, text="🔍", font=("Segoe UI", 8), bg='#007acc', fg='white', relief='flat')
        self.search_button.grid(row=0, column=2, padx=0, sticky="ew")
        clear_button = tk.Button(search_frame, text="Clear", font=("Segoe UI", 8), bg='#d32f2f', fg='white', relief='flat', command=self.clear_search)
        clear_button.grid(row=0, column=3, padx=0, sticky="ew")
        
        sort_frame = tk.Frame(control_frame, bg='#e0e0e0')
        sort_frame.pack(side=tk.RIGHT, padx=10)
        self.sort_label = tk.Label(sort_frame, text=f"Sort {self.device_type} by:", fg="black", bg='#e0e0e0', font=("Segoe UI", 11))
        self.sort_label.pack(side=tk.LEFT, padx=5)
        self.sort_option = tk.StringVar(value=f"{self.device_type} Name")
        self.sort_menu = tk.OptionMenu(sort_frame, self.sort_option, f"{self.device_type} Name", "Response Time", command=self.sort_servers)
        self.sort_menu.config(font=("Segoe UI", 6), bg='#e0e0e0', relief='flat')
        self.sort_menu.pack(side=tk.LEFT)

        # Summary Frame
        summary_frame = tk.Frame(self.root, bg='#1e1e1e')
        summary_frame.pack(fill='x', padx=20, pady=10)
        self.active_count_label = tk.Label(summary_frame, text='Active: 0', fg="#4caf50", font=("Segoe UI", 16, "bold"), bg='#1e1e1e')
        self.active_count_label.pack(side=tk.LEFT, padx=10)
        self.delayed_count_label = tk.Label(summary_frame, text='Delayed: 0', fg="#ff9800", font=("Segoe UI", 16, "bold"), bg='#1e1e1e')
        self.delayed_count_label.pack(side=tk.LEFT, padx=10)
        self.unresponsive_count_label = tk.Label(summary_frame, text='Unresponsive: 0', fg="#f44336", font=("Segoe UI", 16, "bold"), bg='#1e1e1e')
        self.unresponsive_count_label.pack(side=tk.LEFT, padx=10)
        self.analytics_button = tk.Button(summary_frame, text="View Data Analytics", bg='#74a8e7', command=self.select_servers_for_graph, fg="white", font=("Segoe UI", 12), relief="flat")
        self.analytics_button.pack(side=tk.RIGHT, padx=10, pady=10)

        if not self.server_data:
            self.analytics_button.pack_forget()

        # Display Frame
        self.display_frame = tk.Frame(self.root, bg='#1e1e1e')
        self.display_frame.pack(fill='both', expand=True, padx=20, pady=10)

    def read_excel_with_headers(self, file_name, server_column_name, ip_column_name, dashboard_column_name, additional_columns_names):
        wb = openpyxl.load_workbook(filename=file_name, read_only=True)
        ws = wb.active
        self.servers = []

        ### ADDITIONS START
        # If user didn't set self.header_row, fallback to 8
        row_for_headers = (self.header_row - 1) if self.header_row else 8
        ### ADDITIONS END

        # Read the header row (assume it's row 9 in old code, but now can be dynamic)
        header_row = [cell.value for cell in ws[row_for_headers]]

        # Map column names to indices
        header_to_index = {header: idx for idx, header in enumerate(header_row)}

        # Get the indices for the selected columns
        try:
            server_column_idx = header_to_index[server_column_name]
            ip_column_idx = header_to_index[ip_column_name]
            if dashboard_column_name:
                dashboard_column_idx = header_to_index[dashboard_column_name]
            else:
                dashboard_column_idx = None

            # Indices for additional columns
            additional_columns_indices = [header_to_index[col_name] for col_name in additional_columns_names]
        except KeyError as e:
            print(f"Error: Column {e} not found in headers.")
            messagebox.showerror("Error", f"Column {e} not found in headers.")
            return []

        # Now read the data starting from row_for_headers+1 onwards
        start_data_row = row_for_headers + 1  # e.g., if row_for_headers=8, data starts at 9
        for row in ws.iter_rows(min_row=start_data_row):
            server_name = row[server_column_idx].value
            ip_address = row[ip_column_idx].value
            if dashboard_column_idx is not None:
                dashboard_flag = row[dashboard_column_idx].value
            else:
                dashboard_flag = "Yes"

            # Get the values for the additional columns
            additional_values = [row[idx].value for idx in additional_columns_indices]

            if server_name and ip_address and (dashboard_flag == "Yes" or dashboard_column_idx is None):
                self.servers.append([server_name, ip_address, additional_values, None, None, None, None])
                self.response_times[server_name] = []

        wb.close()
        self.filtered_servers = self.servers[:]
        self.additional_columns_names = additional_columns_names
        return self.servers

    def send_custom_notification(self, title, message, server_name, y_offset=0, delay=10):
        """Show custom notification popups using Tkinter at the bottom right of the screen with rounded corners."""
        self.root.after(1, self._create_notification_window, title, message, server_name, y_offset, delay)

    def _create_notification_window(self, title, message, server_name, y_offset, delay):
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()

        notif_width = max(int(screen_width * 0.18), 180)
        notif_height = max(int(screen_height * 0.1), 100)

        title_font_size = max(int(notif_width / 28), 8)
        message_font_size = max(int(notif_width / 35), 5)

        notification_window = Toplevel(self.root)
        notification_window.overrideredirect(1)
        notification_window.attributes("-topmost", True)
        notification_window.wm_attributes("-alpha", 0.0)

        notification_window.geometry(f"{notif_width}x{notif_height}+{screen_width - notif_width - 20}+{screen_height - notif_height - 10 - y_offset}")
        notification_window.wm_attributes("-alpha", 0.95)

        content_frame = tk.Frame(notification_window, bg="white", highlightthickness=0)
        content_frame.place(x=10, y=10, width=notif_width - 20, height=notif_height - 20)

        try:
            icon_image = Image.open("notification_icon.png")
            icon_image = icon_image.resize((int(notif_height * 0.3), int(notif_height * 0.3)))
            icon_photo = ImageTk.PhotoImage(icon_image)
            icon_label = tk.Label(content_frame, image=icon_photo, bg="white")
            icon_label.image = icon_photo
            icon_label.pack(side=tk.LEFT, padx=10, pady=5)
        except Exception as e:
            print(f"Error loading icon: {e}")

        text_frame = tk.Frame(content_frame, bg="white")
        text_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        tk.Label(text_frame, text=title, font=("Segoe UI", title_font_size, "bold"), fg="#434547", bg="white", anchor="w").pack(pady=5, padx=10, anchor="w")
        tk.Label(text_frame, text=message, font=("Segoe UI", message_font_size), fg="#434547", bg="white", anchor="w", justify="left", wraplength=notif_width - 50).pack(pady=5, padx=10, anchor="w")

        notification_window.bind("<Button-1>", lambda event, srv=server_name: self.on_notification_click(srv))

        notification_window.after(delay * 1000, notification_window.destroy)
        self.root.update()

    def on_notification_click(self, server_name):
        if server_name:
            print(f"Notification clicked. Searching for server: {server_name}")
            self.simulate_search(server_name)
        else:
            print("Error: No server name provided.")

    def simulate_search(self, server_name):
        self.search_var.set(server_name.lower())
        self.search_server()

    def notify_server_state_change(self, server_name, server_ip, previous_status, current_status, y_offset):
        message = f"Server: {server_name}\nIP: {server_ip}\n"
        if previous_status == "active" and current_status == "delayed":
            self.send_custom_notification(f"Server Delayed", message + "Status: Delayed", y_offset)
        elif previous_status == "active" and current_status == "unresponsive":
            self.send_custom_notification(f"Server Unresponsive", message + "Status: Unresponsive", y_offset)
        elif previous_status == "delayed" and current_status == "active":
            self.send_custom_notification(f"Server Active", message + "Status: Active", y_offset)
        elif previous_status == "delayed" and current_status == "unresponsive":
            self.send_custom_notification(f"Server Unresponsive", message + "Status: Unresponsive", y_offset)
        elif previous_status == "unresponsive" and current_status == "active":
            self.send_custom_notification(f"Server Active", message + "Status: Active", y_offset)
        elif previous_status == "unresponsive" and current_status == "delayed":
            self.send_custom_notification(f"Server Delayed", message + "Status: Delayed", y_offset)

    def send_notifications(self, active_servers, delayed_servers, unresponsive_servers):
        screen_height = self.root.winfo_screenheight()
        base_y_offset = int(screen_height * 0.11)

        def flatten_server(server):
            if isinstance(server, list) and isinstance(server[0], list):
                server_name = server[0][0]
                server_ip = server[0][1]
            elif isinstance(server, list):
                server_name = server[0]
                server_ip = server[1]
            else:
                server_name = None
                server_ip = None
            return server_name, server_ip

        def send_and_update_offset(title, message, server_name, current_offset):
            self.send_custom_notification(title, message, server_name, current_offset, delay=10)
            return current_offset + base_y_offset

        y_offset = 0
        for server in delayed_servers:
            server_name, server_ip = flatten_server(server)
            if server_name:
                y_offset = send_and_update_offset(
                    f"{self.device_type} Delayed",
                    f"{self.device_type}: {server_name}\nIP: {server_ip}\nStatus: Delayed",
                    server_name,
                    y_offset
                )

        for server in unresponsive_servers:
            server_name, server_ip = flatten_server(server)
            if server_name:
                y_offset = send_and_update_offset(
                    f"{self.device_type} Unresponsive",
                    f"{self.device_type}: {server_name}\nIP: {server_ip}\nStatus: Unresponsive",
                    server_name,
                    y_offset
                )

        for server in active_servers:
            server_name, server_ip = flatten_server(server)
            if server_name:
                if server_name in self.notified_servers["delayed"]:
                    y_offset = send_and_update_offset(
                        f"{self.device_type} Active",
                        f"{self.device_type}: {server_name}\nIP: {server_ip}\nStatus: Active",
                        server_name,
                        y_offset
                    )
                    self.notified_servers["delayed"].discard(server_name)
                    self.notified_servers["active"].add(server_name)

                if server_name in self.notified_servers["unresponsive"]:
                    y_offset = send_and_update_offset(
                        f"{self.device_type} Active",
                        f"{self.device_type}: {server_name}\nIP: {server_ip}\nStatus: Active",
                        server_name,
                        y_offset
                    )
                    self.notified_servers["unresponsive"].discard(server_name)
                    self.notified_servers["active"].add(server_name)

        for server in delayed_servers:
            server_name, server_ip = flatten_server(server)
            if server_name and server_name not in self.notified_servers["delayed"]:
                if server_name in self.notified_servers["active"]:
                    send_and_update_offset(
                        f"{self.device_type} Delayed",
                        f"{self.device_type}: {server_name}\nIP: {server_ip}\nStatus: Delayed",
                        server_name,
                        y_offset
                    )
                    self.notified_servers["active"].discard(server_name)
                self.notified_servers["delayed"].add(server_name)

        for server in unresponsive_servers:
            server_name, server_ip = flatten_server(server)
            if server_name and server_name not in self.notified_servers["unresponsive"]:
                if server_name in self.notified_servers["active"]:
                    send_and_update_offset(
                        f"{self.device_type} Unresponsive",
                        f"{self.device_type}: {server_name}\nIP: {server_ip}\nStatus: Unresponsive",
                        server_name,
                        y_offset
                    )
                    self.notified_servers["active"].discard(server_name)
                    
                elif server_name in self.notified_servers["delayed"]:
                    send_and_update_offset(
                        f"{self.device_type} Unresponsive",
                        f"{self.device_type}: {server_name}\nIP: {server_ip}\nStatus: Unresponsive",
                        server_name,
                        y_offset
                    )
                    self.notified_servers["delayed"].discard(server_name)
                self.notified_servers["unresponsive"].add(server_name)

    async def ping(self, ip, server_name):
        """Ping the server using aioping and return success and response time."""
        try:
            delay = await aioping.ping(ip) * 1000
            self.server_data.setdefault(server_name, []).append(round(delay, 2))
            return True, round(delay, 2)
        except TimeoutError:
            self.server_data.setdefault(server_name, []).append(None)
            return False, None
        except Exception as e:
            print(f"Error pinging {ip}: {e}")
            self.server_data.setdefault(server_name, []).append(None)
            return False, None

    def start_ping(self, servers):
        """Ping all servers asynchronously."""
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        tasks = [self.ping(server[1], server[0]) for server in servers]
        results = loop.run_until_complete(asyncio.gather(*tasks))
        loop.close()

        for index, result in enumerate(results):
            success, response_time = result
            server_name = servers[index][0]
            servers[index][5] = success
            servers[index][6] = response_time

        return servers

    def rerun_aioping(self, reset_cycle=False):
        # Re-read the Excel file data if available
        if hasattr(self, 'excel_file'):
            print("Re-reading Excel file data...")
            new_data = self.read_excel_with_headers(
                self.excel_file,
                self.server_column_name,
                self.ip_column_name,
                self.dashboard_column_name,
                self.additional_columns_names
            )
            if new_data:
                self.servers = new_data

        if self.ping_in_progress:
            print("Ping in progress... skipping this cycle.")
            return

        self.ping_in_progress = True
        print("Starting a new ping cycle...")

        ping_results = self.start_ping(self.servers)

        active_servers = [server for server in self.servers if server[5] and server[6] <= 200]
        delayed_servers = [server for server in self.servers if server[5] and server[6] > 200]
        unresponsive_servers = [server for server in self.servers if not server[5]]

        if self.sort_requested:
            print(f"Applying sorting by: {self.sort_by}")
            self.sort_requested = False
            
            if self.sort_by == f"{self.device_type} Name":
                self.servers.sort(key=lambda x: x[0].strip() if isinstance(x[0], str) else x[0])
            elif self.sort_by == "Response Time":
                self.servers.sort(key=lambda x: (x[6] if x[6] is not None else float('inf')))

        self.categorize_servers(self.servers)
        self.send_notifications(active_servers, delayed_servers, unresponsive_servers)
        
        if reset_cycle:
            self.current_page_index = 0
            self.cycle_pages()

        self.calculate_ping_interval()

        print("Ping cycle complete.")
        self.ping_in_progress = False
        self.schedule_next_ping()

    def calculate_ping_interval(self):
        total_pages = sum(len(self.pages[category]) for category in self.pages)
        full_cycle_time = total_pages * self.cycle_interval

        min_ping_interval_in_seconds = full_cycle_time // 1000

        if not self.ping_interval_manually_changed:
            self.ping_interval_var.set(min_ping_interval_in_seconds)
            self.ping_interval = min_ping_interval_in_seconds * 1000
        else:
            self.ping_interval = self.ping_interval_var.get() * 1000

        print(f"Ping interval set to {self.ping_interval / 1000} seconds (Calculated or Manual).")

    def schedule_next_ping(self):
        interval_in_ms = self.ping_interval if self.ping_interval is not None else 10000
        print(f"Scheduling next ping in {interval_in_ms / 1000} seconds.")

        self.ping_cycle_id = self.root.after(interval_in_ms, self.rerun_aioping)
        self.ping_in_progress = False

    def cancel_ping_cycle(self):
        if hasattr(self, 'ping_cycle_id'):
            self.root.after_cancel(self.ping_cycle_id)
            print("Canceled existing ping cycle.")
        if hasattr(self, 'page_cycle_id'):
            self.root.after_cancel(self.page_cycle_id)
            print("Canceled page cycling.")
        self.ping_in_progress = False

    def categorize_servers(self, servers):
        active = []
        delayed = []
        unresponsive = []

        for server in servers:
            if server[5] and server[6] is not None and server[6] <= 200:
                active.append(server)
            elif server[5]:
                delayed.append(server)
            else:
                unresponsive.append(server)

        self.pages["Active"] = self.paginate(active)
        self.pages["Delayed"] = self.paginate(delayed)
        self.pages["Unresponsive"] = self.paginate(unresponsive)
        
        self.update_summary(len(active), len(delayed), len(unresponsive))

    def toggle_auto_refresh(self):
        if self.auto_refresh_var.get():
            print("Auto-refresh turned on.")
            self.cancel_ping_cycle()
            self.current_page_index = 0
            self.rerun_aioping(reset_cycle=True)
            self.cycle_pages()
        else:
            print("Auto-refresh turned off.")
            self.cancel_ping_cycle()

    def display_page(self, data, category):
        for widget in self.display_frame.winfo_children():
            widget.destroy()

        display_width = self.root.winfo_width()
        display_height = self.root.winfo_height() - 200

        columns = 5
        rows = 5
        card_width = display_width // columns - 20
        card_height = display_height // rows - 20

        name_font_size = max(int(card_width / 18), 10)
        ip_font_size = max(int(card_width / 22), 8)
        response_font_size = max(int(card_width / 22), 8)

        data = data[:25]

        bg_color = "#4caf50" if category == "Active" else "#ff9800" if category == "Delayed" else "#f44336"

        for i, items in enumerate(data):
            if items:
                row_frame = tk.Frame(self.display_frame, bg='#1e1e1e')
                row_frame.grid(row=i // columns, column=i % columns, padx=10, pady=10, sticky="nsew")

                card = tk.Frame(row_frame, bg=bg_color, bd=2, relief="groove", width=card_width, height=card_height)
                card.pack_propagate(False)
                card.pack(fill="both", expand=True)

                card.bind("<Button-1>", lambda event, srv=items, cw=card_width, ch=card_height: self.on_card_click(event, srv, cw, ch))

                name_label = tk.Label(card, text=f"{items[0]}", fg="white", font=("Segoe UI", name_font_size, "bold"), bg=bg_color)
                name_label.pack(side=tk.TOP, anchor="w", padx=10, pady=5)
                name_label.bind("<Button-1>", lambda event, srv=items, cw=card_width, ch=card_height: self.on_card_click(event, srv, cw, ch))

                ip_label = tk.Label(card, text=f"IP: {items[1]}", fg="white", font=("Segoe UI", ip_font_size), bg=bg_color)
                ip_label.pack(side=tk.TOP, anchor="w", padx=10)
                ip_label.bind("<Button-1>", lambda event, srv=items, cw=card_width, ch=card_height: self.on_card_click(event, srv, cw, ch))

                response_time = f"Response: {items[6]} ms" if items[5] else "Response: No Response"
                response_label = tk.Label(card, text=response_time, fg="white", font=("Segoe UI", response_font_size), bg=bg_color)
                response_label.pack(side=tk.TOP, anchor="w", padx=10, pady=2)
                response_label.bind("<Button-1>", lambda event, srv=items, cw=card_width, ch=card_height: self.on_card_click(event, srv, cw, ch))

        if data:
            self.analytics_button.pack(side=tk.RIGHT, padx=10, pady=10)
        else:
            self.analytics_button.pack_forget()

    def on_card_click(self, event, server, card_width, card_height):
        self.auto_refresh_var.set(False)
        self.auto_refresh_checkbox.deselect()
        self.cancel_ping_cycle()
        self.enlarge_card(server, card_width, card_height)

    def enlarge_card(self, server, original_card_width, original_card_height):
        for widget in self.display_frame.winfo_children():
            widget.destroy()

        card_width = int(original_card_width * 2.2)
        card_height = int(original_card_height * 2.2)

        self.display_frame.update_idletasks()
        display_width = self.display_frame.winfo_width()
        display_height = self.display_frame.winfo_height()

        self.enlarged_frame = tk.Frame(self.display_frame, bg='#1e1e1e')
        x_position = (display_width - card_width) // 2
        y_position = (display_height - card_height) // 2
        self.enlarged_frame.place(x=x_position, y=y_position)

        self.card_canvas = tk.Canvas(self.enlarged_frame, width=card_width, height=card_height, bg='#1e1e1e', highlightthickness=0)
        self.card_canvas.pack()

        self.prepare_card_images(server, card_width, card_height)

        self.current_side = 'front'
        self.card_image_id = self.card_canvas.create_image(card_width / 2, card_height / 2, image=self.front_image_tk, anchor='center')

        self.card_canvas.tag_bind(self.card_image_id, "<Button-1>", lambda event: self.flip_card_canvas(event, server))

        self.auto_flipped = False
        self.root.after(200, self.auto_flip_card)

    def auto_flip_card(self):
        if not self.auto_flipped:
            self.auto_flipped = True
            self.flip_card_canvas(None, auto_flip=True)

    def close_enlarged_card(self):
        if self.flip_animation_running:
            print("Animation in progress, cannot close the card.")
            return

        self.enlarged_frame.destroy()
        self.auto_flipped = False
        self.auto_refresh_var.set(True)
        self.auto_refresh_checkbox.select()
        self.cancel_ping_cycle()
        self.current_page_index = 0
        self.rerun_aioping(reset_cycle=True)
        self.cycle_pages()

    def prepare_card_images(self, server, card_width, card_height):
        self.front_image = self.create_card_image(server, card_width, card_height, side='front')
        self.back_image = self.create_card_image(server, card_width, card_height, side='back')

        self.front_image_tk = ImageTk.PhotoImage(self.front_image)
        self.back_image_tk = ImageTk.PhotoImage(self.back_image)

    def create_card_image(self, server, card_width, card_height, side='front'):
        bg_color = "#4caf50" if server[5] and server[6] <= 200 else "#ff9800" if server[5] else "#f44336"

        if side == 'front':
            background_color = bg_color
        else:
            background_color = "white"

        image = Image.new('RGBA', (card_width, card_height), background_color)
        draw = ImageDraw.Draw(image)

        if side == 'back':
            border_width = 4
            draw.rectangle(
                [(border_width // 2, border_width // 2), (card_width - border_width // 2 - 1, card_height - border_width // 2 - 1)],
                outline=bg_color, width=border_width
            )

        name_font_size = max(int(card_width / 15), 12)
        detail_font_size = max(int(card_width / 20), 10)
        more_detail_font_size = max(int(card_width / 25), 8)
        smaller_bracket_font_size = more_detail_font_size - 4
        close_button_font_size = max(int(card_width / 30), 6)

        try:
            close_button_font = ImageFont.truetype("seguisym.ttf", close_button_font_size)
            font = ImageFont.truetype("seguisym.ttf", name_font_size)
            detail_font = ImageFont.truetype("seguisym.ttf", detail_font_size)
            more_detail_font = ImageFont.truetype("seguisym.ttf", more_detail_font_size)
            smaller_bracket_font = ImageFont.truetype("seguisym.ttf", smaller_bracket_font_size)
        except IOError:
            close_button_font = ImageFont.load_default()
            font = ImageFont.load_default()
            detail_font = ImageFont.load_default()
            more_detail_font = ImageFont.load_default()
            smaller_bracket_font = ImageFont.load_default()

        if background_color == 'white':
            close_button_fill_color = bg_color
        else:
            close_button_fill_color = 'white'

        close_button_text = "✖"
        bbox = close_button_font.getbbox(close_button_text)
        text_width = bbox[2] - bbox[0]
        text_height = bbox[3] - bbox[1]

        draw.text(
            (card_width - text_width - 10, 10),
            close_button_text,
            font=close_button_font,
            fill='white'
        )

        self.close_button_text_width = text_width
        self.close_button_text_height = text_height

        if side == 'front':
            text_color = 'white'
            draw.text((20, 20), f"{server[0]}", font=font, fill=text_color)
            draw.text((20, 100), f"IP: {server[1]}", font=detail_font, fill=text_color)
            response_time = f"Response: {server[6]} ms" if server[5] else "Response: No Response"
            draw.text((20, 150), response_time, font=detail_font, fill=text_color)
        else:
            text_color = 'black'
            draw.text((20, 20), f"{self.device_type} Details:", font=font, fill=text_color)
            y_position = 80
            line_spacing = more_detail_font_size + 5

            max_line_width = card_width - 40

            fonts = {'normal': more_detail_font, 'small': smaller_bracket_font}
            
            def wrap_text(text, font, max_width):
                words = text.split()
                lines = []
                current_line = ''
                for word in words:
                    test_line = current_line + (' ' if current_line else '') + word
                    bbox = font.getbbox(test_line)
                    test_width = bbox[2] - bbox[0]
                    if test_width <= max_width:
                        current_line = test_line
                    else:
                        if current_line:
                            lines.append(current_line)
                        current_line = word
                if current_line:
                    lines.append(current_line)
                return lines

            for col_name, value in zip(self.additional_columns_names, server[2]):
                value = "N/A" if value is None else str(value)
                value_lines = value.split('\n')

                for idx, val_line in enumerate(value_lines):
                    line_text = f"{col_name}: {val_line}" if idx == 0 else val_line
                    wrapped_lines = wrap_text(line_text, more_detail_font, max_line_width)
                    for wrapped_line in wrapped_lines:
                        draw.text((20, y_position), wrapped_line, font=more_detail_font, fill=text_color)
                        y_position += line_spacing
                y_position += 5

            draw.text((20, y_position + 5), f"Last Checked: {time.strftime('%d/%m/%Y %H:%M:%S')}", font=more_detail_font, fill=text_color)

        return image

    def is_click_on_close_button(self, x, y):
        """Check if the click coordinates are within the close button area."""
        # Expand the clickable bounding box by some margin (say 10px)
        margin = 10

        close_button_x1 = self.original_width - self.close_button_text_width - margin
        close_button_y1 = 10 - margin
        close_button_x2 = self.original_width - 10 + margin
        close_button_y2 = 10 + self.close_button_text_height + margin

        return (close_button_x1 <= x <= close_button_x2) and (close_button_y1 <= y <= close_button_y2)

    def flip_card_canvas(self, event, server=None, auto_flip=False):
        self.original_width = self.card_canvas.winfo_width()
        self.original_height = self.card_canvas.winfo_height()

        if event and self.is_click_on_close_button(event.x, event.y):
            self.close_enlarged_card()
            return

        if self.flip_animation_running:
            return

        self.flip_animation_running = True
        self.steps = 40
        self.delay = 10

        self.animate_flip(0)
    
    def animate_flip(self, step):
        if not self.card_canvas.winfo_exists():
            print("Canvas has been destroyed, stopping animation.")
            self.flip_animation_running = False
            return

        if step <= self.steps:
            scale = math.cos((step / self.steps) * (math.pi / 2))
            scale = max(scale, 0.01)
            self.card_canvas.scale("all", self.original_width / 2, self.original_height / 2, scale, 1)
            
            alpha = int(255 * scale)
            if self.current_side == 'front':
                image = self.front_image.copy()
            else:
                image = self.back_image.copy()
            image.putalpha(alpha)
            image_tk = ImageTk.PhotoImage(image)
            self.card_canvas.itemconfig(self.card_image_id, image=image_tk)
            self.card_canvas.image = image_tk

            self.card_canvas.update()
            self.root.after(self.delay, lambda: self.animate_flip(step + 1))
        else:
            self.current_side = 'back' if self.current_side == 'front' else 'front'
            self.animate_unflip(1)

    def animate_unflip(self, step):
        """Animate the unflip of the card back to its original state."""
        if not self.card_canvas.winfo_exists():
            print("Canvas has been destroyed, stopping animation.")
            self.flip_animation_running = False  # Reset flag if the canvas is destroyed
            return

        if step <= self.steps:
            scale = math.sin((step / self.steps) * (math.pi / 2))
            scale = max(scale, 0.01)  # Prevent scale from being zero
            self.card_canvas.scale("all", self.original_width / 2, self.original_height / 2, scale, 1)
            
            # Apply fade effect
            alpha = int(255 * scale)
            if self.current_side == 'front':
                image = self.front_image.copy()
            else:
                image = self.back_image.copy()
            image.putalpha(alpha)
            image_tk = ImageTk.PhotoImage(image)
            self.card_canvas.itemconfig(self.card_image_id, image=image_tk)
            self.card_canvas.image = image_tk  # Keep a reference

            self.card_canvas.update()
            self.root.after(self.delay, lambda: self.animate_unflip(step + 1))
        else:
            # Reset the scale to original in case of any rounding errors
            self.card_canvas.scale("all", self.original_width / 2, self.original_height / 2, 1, 1)

            # Reset the flip animation flag once the unflip completes
            self.flip_animation_running = False

    def update_summary(self, active_count, delayed_count, unresponsive_count):
        """Update the category summary with the number of active, delayed, and unresponsive servers."""
        self.active_count_label.config(text=f'Active: {active_count}')
        self.delayed_count_label.config(text=f'Delayed: {delayed_count}')
        self.unresponsive_count_label.config(text=f'Unresponsive: {unresponsive_count}')

    def paginate(self, servers):
        """Paginate the list of servers."""
        return [servers[i:i + self.max_cards_per_page] for i in range(0, len(servers), self.max_cards_per_page)]

    def cycle_pages(self):
        """Cycle through server pages, ensuring pages are displayed independently of pinging."""
        if not self.auto_refresh_var.get():
            print("Auto-refresh is off. Stopping page cycle.")
            return  # Stop cycling pages if auto-refresh is off

        if self.ping_in_progress:
            print("Ping in progress. Pausing page cycling.")
            return  # Stop cycling pages if a ping is in progress

        total_pages = sum(len(self.pages[category]) for category in self.pages)

        # Determine the current category and page to display
        accumulated_pages = 0
        current_category_index = 0
        for i, category in enumerate(self.categories):
            if accumulated_pages + len(self.pages[category]) > self.current_page_index:
                current_category_index = i
                break
            accumulated_pages += len(self.pages[category])

        current_category = self.categories[current_category_index]
        category_pages = self.pages[current_category]
        page_number = self.current_page_index - accumulated_pages

        if page_number < len(category_pages):
            # Display the current page
            self.display_page(category_pages[page_number], current_category)
        else:
            self.current_page_index += 1
            if self.current_page_index >= total_pages:
                self.current_page_index = 0
            self.cycle_pages()
            return

        # Move to the next page
        self.current_page_index += 1
        if self.current_page_index >= total_pages:
            self.current_page_index = 0

        # Continue cycling every 5 seconds
        if self.auto_refresh_var.get() and not self.ping_in_progress:
            print("Cycling to the next page.")
            self.page_cycle_id = self.root.after(self.cycle_interval, self.cycle_pages)

    def search_server(self):
        """Search for a server by name or IP."""
        search_term = self.search_var.get().lower()
        self.filtered_servers = [server for server in self.servers if search_term in server[0].lower() or search_term in server[1].lower()]

        if self.filtered_servers:
            self.auto_refresh_var.set(False)  # Disable auto-refresh on the first search
            self.display_search_result(self.filtered_servers[0])
            self.cancel_ping_cycle()  # Stop the page cycling

    def clear_search(self):
        """Clear the search input and restore the full server list."""
        self.search_var.set('')  # Clear the search entry
        self.filtered_servers = self.servers[:]  # Reset to the full server list
        self.search_active = False

    def display_search_result(self, server):
        for widget in self.display_frame.winfo_children():
            widget.destroy()

        # Get the available display area (excluding some margin for top/bottom controls)
        display_width = self.root.winfo_width()
        display_height = self.root.winfo_height() - 200  # Subtract space for top/bottom margins

        # Define the number of rows and columns for displaying 25 cards
        columns = 5
        rows = 5
        card_width = display_width // columns - 20  # Account for padding/margin
        card_height = display_height // rows - 20

        # Font size based on card size (adjust dynamically)
        title_font_size = max(int(card_width / 18), 10)
        content_font_size = max(int(card_width / 22), 8)

        row_frame = tk.Frame(self.display_frame, bg='#1e1e1e')
        row_frame.grid(row=0, column=0, padx=10, pady=10, sticky="nsew")

        # Determine the background color based on server status
        if server[5]:  # If the server responded
            if server[6] <= 200:  # Response time less than or equal to 200ms -> Active
                bg_color = "#4caf50"  # Green for Active
            else:
                bg_color = "#ff9800"  # Orange for Delayed (response time more than 200ms)
        else:
            bg_color = "#f44336"  # Red for Unresponsive
            
        card = tk.Frame(row_frame, bg=bg_color, bd=2, relief="groove", height=card_height, width=card_width)
        card.pack_propagate(False)
        card.pack(fill="both", expand=True)

        # Bind click event to the card and labels
        card.bind("<Button-1>", lambda event: self.enlarge_search_result_card(event, server, card_width, card_height))

        # Server name
        name_label = tk.Label(card, text=f"{server[0]}", fg="white", font=("Segoe UI", title_font_size, "bold"), bg=bg_color)
        name_label.pack(side=tk.TOP, anchor="w", padx=10, pady=5)
        name_label.bind("<Button-1>", lambda event: self.enlarge_search_result_card(event, server, card_width, card_height))

        # Server IP
        ip_label = tk.Label(card, text=f"IP: {server[1]}", fg="white", font=("Segoe UI", content_font_size), bg=bg_color)
        ip_label.pack(side=tk.TOP, anchor="w", padx=10)
        ip_label.bind("<Button-1>", lambda event: self.enlarge_search_result_card(event, server, card_width, card_height))

        # Response time
        response_time = f"Response: {server[6]} ms" if server[5] else "Response: No Response"
        response_label = tk.Label(card, text=response_time, fg="white", font=("Segoe UI", content_font_size), bg=bg_color)
        response_label.pack(side=tk.TOP, anchor="w", padx=10, pady=2)
        response_label.bind("<Button-1>", lambda event: self.enlarge_search_result_card(event, server, card_width, card_height))

        if self.auto_refresh_var.get():
            self.auto_refresh_var.set(False)
            self.toggle_auto_refresh()
        
        if server:
            self.analytics_button.pack(side=tk.RIGHT, padx=10, pady=10)  # Show the button if there is server data
        else:
            self.analytics_button.pack_forget()  # Hide the button if no data is available

    def enlarge_search_result_card(self, event, server, original_card_width, original_card_height):
        """Enlarge the search result card and apply the flip animation."""
        # Clear the display frame
        for widget in self.display_frame.winfo_children():
            widget.destroy()

        # Calculate the enlarged card size as 30% bigger than the original size
        card_width = int(original_card_width * 2.2)
        card_height = int(original_card_height * 2.2)

        # Get the dimensions of the display_frame
        self.display_frame.update_idletasks()  # Ensure dimensions are updated
        display_width = self.display_frame.winfo_width()
        display_height = self.display_frame.winfo_height()

        # Create a new frame for the enlarged card
        self.enlarged_frame = tk.Frame(self.display_frame, bg='#1e1e1e')
        # Center the enlarged card on the dashboard
        x_position = (display_width - card_width) // 2
        y_position = (display_height - card_height) // 2
        self.enlarged_frame.place(x=x_position, y=y_position)

        # Create a Canvas for advanced animations
        self.card_canvas = tk.Canvas(self.enlarged_frame, width=card_width, height=card_height, bg='#1e1e1e', highlightthickness=0)
        self.card_canvas.pack()

        # Prepare front and back images of the card
        self.prepare_card_images(server, card_width, card_height)

        # Start with the front side displayed
        self.current_side = 'front'
        self.card_image_id = self.card_canvas.create_image(card_width / 2, card_height / 2, image=self.front_image_tk, anchor='center')

        # Bind click event to flip the card
        self.card_canvas.tag_bind(self.card_image_id, "<Button-1>", lambda event: self.flip_card_canvas(event, server))

    def sort_servers(self, option):
        """Mark that sorting has been requested, to be applied after the next reping."""
        self.sort_by = option
        self.sort_requested = True  # Sorting will be applied after the next ping cycle
        print(f"Sorting requested by: {self.sort_by}. Sorting will take effect after the next ping cycle.")

    def select_servers_for_graph(self):
        """Dialog box for selecting specific servers to view on the graph.""" 
        if not self.server_data:
            messagebox.showerror("No Data", "No response data available for plotting.")
            return

        # Create a new dialog window for server selection
        dialog = tk.Toplevel(self.root)
        dialog.geometry("400x350")
        dialog.title(f"Select {self.device_type} to Plot")
        dialog.configure(bg='#2b2b2b')

        # Create a listbox with a scrollbar for server selection
        listbox = tk.Listbox(dialog, selectmode=tk.MULTIPLE, font=("Segoe UI", 10), fg="white", bg='#2b2b2b')
        scrollbar = tk.Scrollbar(dialog)
        listbox.config(yscrollcommand=scrollbar.set)
        scrollbar.config(command=listbox.yview)
        listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=10, pady=10)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        # Insert server names into the listbox
        for server_name in self.server_data.keys():
            listbox.insert(tk.END, server_name)

        # Create a frame for the plot button
        button_frame = tk.Frame(dialog, bg='#2b2b2b')
        button_frame.pack(side=tk.BOTTOM, fill=tk.X, pady=10)

        # Function to plot the selected servers
        def plot_selected():
            selected_indices = listbox.curselection()  # Get selected indices
            selected_servers = [listbox.get(i) for i in selected_indices]  # Get server names from indices
            if selected_servers:
                self.plot_graph(selected_servers)  # Plot the selected servers
            else:
                messagebox.showerror("No Selection", "Please select at least one server to plot.")
            dialog.destroy()

        # Add the "Plot Selected Servers" button at the bottom
        plot_button = tk.Button(button_frame, text=f"Plot Selected {self.device_type}", command=plot_selected,
                                bg='#74a8e7', fg="white", font=("Segoe UI", 11), relief="flat")
        plot_button.pack(pady=140, padx=5)

    def plot_graph(self, selected_servers):
        """Plot the response times of selected servers."""
        if hasattr(self, 'graph_window') and self.graph_window.winfo_exists():
            self.graph_window.destroy()
        
        fig, ax = plt.subplots(figsize=(8, 6))

        # Plot the response time data for each selected server
        for server_name in selected_servers:
            response_times = self.server_data.get(server_name, [])

            filtered_response_times = [r for r in response_times if r is not None]
            
            # Ensure there is data to plot
            if filtered_response_times and len(filtered_response_times) > 1:
                print(f"Plotting data for {server_name}: {response_times}")
                x_vals = range(1, len(response_times) + 1)  # Dynamically set x-axis based on actual number of pings
                y_vals = response_times
                ax.plot(x_vals, y_vals, label=server_name, linewidth=1)

                # Find the highest point and annotate it
                max_response_time = max(filtered_response_times)
                max_index = response_times.index(max_response_time) + 1  # +1 because x_vals starts at 1
                ax.annotate(f'{max_response_time}', xy=(max_index, max_response_time), xytext=(max_index, max_response_time+5),
                            arrowprops=dict(facecolor='black', shrink=0.05), fontsize=8)
            else:
                print(f"No valid data to plot for {server_name}")

        # Add the 200ms dotted line
        ax.axhline(y=200, color='r', linestyle='--', label='Delayed Threshold (200 ms)')

        # Set the graph title and labels
        ax.set_title(f"{self.device_type} Response Time Analytics")
        ax.set_xlabel('Ping Attempt')
        ax.set_ylabel('Response Time (ms)')

        # Adjust x-axis limits dynamically based on number of ping attempts
        max_attempts = max([len(self.server_data.get(server_name, [])) for server_name in selected_servers], default=1)
        if max_attempts > 1:
            ax.set_xlim(1, max_attempts)  # Set x-axis to number of attempts
        else:
            ax.set_xlim(0, 2)  # Prevent issue with single-point graphs

        # Add the legend outside the plot, to avoid cluttering the graph itself
        ax.legend(loc='upper right', fontsize=7)

        # Show the plot in the Tkinter window
        self.graph_window = tk.Toplevel(self.root)
        self.graph_window.title("Response Time Analytics")
        
        # Keep the window on top and prevent accidental closure on repinging
        self.graph_window.wm_attributes("-topmost", True)
        self.graph_window.focus_force()
        self.graph_window.protocol("WM_DELETE_WINDOW", self.graph_window.destroy)

        canvas = FigureCanvasTkAgg(fig, master=self.graph_window)
        canvas.draw()
        canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

    def on_slider_change(self, event=None):
        """Handle changes to the ping interval slider and set the ping interval manually."""
        new_interval_in_seconds = self.ping_interval_var.get()
        
        # Set the new ping interval
        self.ping_interval = new_interval_in_seconds * 1000  # Convert to milliseconds
        self.ping_interval_manually_changed = True  # Mark that the slider has been manually adjusted
        print(f"Ping interval manually changed to {new_interval_in_seconds} seconds.")
        
        # Restart the ping cycle with the new interval
        self.rerun_aioping(reset_cycle=True)

    def exit_app(self):
        self.root.attributes("-fullscreen", False)
        self.root.destroy()


if __name__ == "__main__":
    root = tk.Tk()
    app = PingingDashboard(root)
    root.mainloop()
